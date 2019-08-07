import openpyxl
import random

print("Hello, please wait for checking...")

filename = "receipt_example.xlsx"
book = openpyxl.load_workbook(filename)
sht_receipt = book['receipt']
sht_grade = book['grade']
sht_menu = book['menu']
sht_branch = book['branch']
sht_order_type = book['order_type']
sht_size = book['size']
sht_age = book['age']

print("---sheet import done.---")

def amount():
    sht_receipt.cell(row = ID, column = 6, value = '1')

def time():
    hour = str(random.randrange(9,23))
    minute = str(random.randrange(00,60))
    sht_receipt.cell(row = ID, column = 2, value = hour + ':' + minute)

def age():
    age_from = int(sht_age.cell(row = 1, column = 1).value)
    age_to = int(sht_age.cell(row = 2, column = 1).value) + 1
    sht_receipt.cell(row = ID, column = 9, value = str(random.randrange(age_from, age_to)))

def sex():
    sht_receipt.cell(row = ID, column = 10, value = str(random.choice(['Male', 'Female'])))

def take_out():
    sht_receipt.cell(row = ID, column = 11, value = str(random.choice(['Yes', 'No'])))

def tumblr():
    sht_receipt.cell(row = ID, column = 12, value = str(random.choice(['Yes', 'No'])))

def get_num(sht_):
    num = 0
    for r in sht_:
        num += 1
    return num + 1

def grade(i, grade_n):
    pick_grade = random.randrange(1, grade_n)
    sht_receipt.cell(row = i, column = 8, value = sht_grade.cell(row = pick_grade, column = 1).value)

def branch(i, branch_n):
    pick_branch = random.randrange(1, branch_n)
    sht_receipt.cell(row = i, column = 14, value = sht_branch.cell(row = pick_branch, column = 1).value)

def order_type(i, order_type_n):
    pick_order_type = random.randrange(1, order_type_n)
    sht_receipt.cell(row = i, column = 13, value = sht_order_type.cell(row = pick_order_type, column = 1).value)

def size(i, size_n):
    pick_size = random.randrange(1, size_n)
    sht_receipt.cell(row = i, column = 5, value = sht_size.cell(row = pick_size, column = 1).value)

def menu(i, menu_num, size_n):
    pick_menu = random.randrange(2, menu_num)
    sht_receipt.cell(row = i, column = 3, value = sht_menu.cell(row = pick_menu, column = 1).value)
    sht_receipt.cell(row = i, column = 7, value = sht_menu.cell(row = pick_menu, column = 2).value)
    sht_receipt.cell(row = i, column = 4, value = sht_menu.cell(row = pick_menu, column = 3).value)
    if  (sht_menu.cell(row = pick_menu, column = 3).value == 'coffee'):
        size(i, size_n)
    else:
        sht_receipt.cell(row = i, column = 5, value = 'N/A')

menu_n = get_num(sht_menu.rows)
print("--> get number of 'menu': {0} done.".format(menu_n - 1))
grade_n = get_num(sht_grade.rows)
print("--> get number of 'grade': {0} done.".format(grade_n - 1))
branch_n = get_num(sht_branch.rows)
print("--> get number of 'branch': {0} done.".format(branch_n - 1))
order_type_n = get_num(sht_order_type.rows)
print("--> get number of 'order_type': {0} done.".format(order_type_n - 1))
size_n = get_num(sht_size.rows)
print("--> get number of 'size': {0} done.".format(size_n - 1))
print("--- initializing done ---")
print('')

ID = 2
print(">>> Put number of ID: "),
ID_num = input()
while ID < int(ID_num) + 2:
    sht_receipt.cell(row = ID, column = 1, value = ID - 1)

    menu(ID, menu_n, size_n)
    grade(ID, grade_n)
    order_type(ID, order_type_n)
    branch(ID, branch_n)
    time()
    age()
    sex()
    take_out()
    tumblr()
    amount()

    print(str(ID - 1) + " done")
    ID += 1

book.save(filename)
print("--> saving excel file done.")
print("--good--")